import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.styles import Font

outputPath = '/Users/macbook/Desktop/SkinGPT-X-EvaluationResults/DDI/output/'
dataPath   = '/Users/macbook/Desktop/SkinGPT-X-Dataset/DDI/'
df = pd.read_csv(outputPath + 'merged_diagnostic_assessment_with_groundtruth_clean.csv')

wb = Workbook()
ws = wb.active
ws.title = "Results"

# 1. 表头
for col_idx, col_name in enumerate(df.columns, 1):
    ws.cell(row=1, column=col_idx, value=col_name)

red_font = Font(color="FF0000")

# 2. 逐行写入
for row_idx, row in df.iterrows():
    groundtruth = str(row['groundtruth']).strip().lower()
    keywords = [w.strip() for w in groundtruth.split('-') if w.strip()]

    # 2.1 固定列：image_name & groundtruth
    ws.cell(row=row_idx + 2, column=1, value=str(row['image_name']))
    ws.cell(row=row_idx + 2, column=5, value=str(row['groundtruth']))

    # 2.2 需要比对的列：B/C/D/F
    for col_offset, col_name in enumerate(['RAG_diagnostic_assessment',
                                         'SkinGPT_diagnostic_assessment',
                                         'WebSearch_diagnostic_assessment',
                                         'PrimaryDiagnosis'], start=2):
        original = str(row[col_name])
        cell = ws.cell(row=row_idx + 2, column=col_offset, value=original)

        if keywords:
            pattern = re.compile(r'\b(' + '|'.join(map(re.escape, keywords)) + r')\b',
                               flags=re.IGNORECASE)
            if pattern.search(original):
                cell.font = red_font  # 整格标红

# 3. 保存
wb.save(outputPath + "output_highlighted.xlsx")
print("处理完成（含 PrimaryDiagnosis 标红），已保存为 output_highlighted.xlsx")