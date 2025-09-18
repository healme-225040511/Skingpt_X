#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
All-in-one 诊断结果后处理工具
--------------------------------
prepare_csv()      -> 合并 4 个 agent CSV + groundtruth + JSON PrimaryDiagnosis
clean_nan()        -> 去掉空值行
extract_assessment() -> 从任意 agent JSON 提取 DiagnosticAssessment
extract_treatment()  -> 专用于 TreatmentRecommend_output.json 提取 PrimaryDiagnosis
highlight_excel()  -> 生成关键词高亮 Excel
--------------------------------
命令行示例（与老脚本完全兼容）：
# 1. 合并
python all_in_one_diagnosis_utils.py prepare_csv
# 2. 清洗
python all_in_one_diagnosis_utils.py clean_nan
# 3. 提取（通用）
python all_in_one_diagnosis_utils.py extract_assessment --json output/RAG_output.json --agent_name RAG
# 4. 提取（Treatment 专用）
python all_in_one_diagnosis_utils.py extract_treatment
# 5. 高亮
python all_in_one_diagnosis_utils.py highlight_excel
"""
from __future__ import annotations
import json
import csv
import re
import argparse
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font
import pandas as pd
import json
import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def safe_get_primary_diagnosis(data, img_name):
    val = data.get(img_name, "")
    if isinstance(val, dict):
        return val.get("Primary Diagnosis", "")
    elif isinstance(val, str):
        match = re.search(r"Primary Diagnosis[:\s]*([^\n]*)", val, flags=re.I)
        return match.group(1).strip() if match else val.strip()
    return ""


def ClassificationResultSaveToCSV(RAG_output, SkinGPT_output, WebSearch_output, CaseReview_output, Reasoning_output,
                                  TreatmentRecommend_output, output_csv_path):
    # 定义JSON文件的路径
    file_paths = {
        "RAG_output": RAG_output,
        "SkinGPT_output": SkinGPT_output,
        "WebSearch_output": WebSearch_output,
        "CaseReview_output": CaseReview_output,
        "Reasoning_output": Reasoning_output,
        "TreatmentRecommend_output": TreatmentRecommend_output,
    }

    # 输出CSV文件的路径
    output_csv_path = output_csv_path

    # 函数：加载JSON文件内容
    def load_json_file(filepath):
        """
        加载指定路径的JSON文件。如果文件不存在，则打印警告并返回空字典。
        """
        if not os.path.exists(filepath):
            print(f"警告：文件未找到：'{filepath}'。将跳过此文件。")
            return {}
        try:
            with open(filepath, 'r', encoding='utf-8') as f:
                return json.load(f)
        except json.JSONDecodeError:
            print(f"警告：文件 '{filepath}' 的JSON解码失败。可能为空或格式错误。将返回空字典。")
            return {}
        except Exception as e:
            print(f"读取文件 '{filepath}' 时发生错误: {e}")
            return {}

    # 加载所有JSON数据
    all_data = {key: load_json_file(path) for key, path in file_paths.items()}

    # 收集所有唯一的图像文件名
    all_image_names = set()
    for data_dict in all_data.values():
        all_image_names.update(data_dict.keys())

    # 准备用于DataFrame的数据列表
    csv_rows = []

    # 遍历所有收集到的图像文件名，提取对应的值
    for img_name in sorted(list(all_image_names)):
        row = {"图像名": img_name}

        # 从 RAG_output, SkinGPT_output, WebSearch_output 中提取 'Diagnostic Assessment' 后的值
        # 这些文件中的"Diagnostic Assessment"键对应的值是一整个长的字符串，按要求直接提取该字符串
        row["RAG_output_诊断评估值"] = safe_get_primary_diagnosis(all_data["RAG_output"], img_name)
        row["SkinGPT_output_诊断评估值"] = safe_get_primary_diagnosis(all_data["SkinGPT_output"], img_name)
        row["WebSearch_output_诊断评估值"] = safe_get_primary_diagnosis(all_data["WebSearch_output"], img_name)

        # 从 CaseReview_output, Reasoning_output, TreatmentRecommend_output 中提取 'PrimaryDiagnosis' 的值
        # 这些文件中的"PrimaryDiagnosis"键对应的值通常是直接的诊断字符串
        row["CaseReview_output_PrimaryDiagnosis"] = all_data["CaseReview_output"].get(img_name, {}).get("PrimaryDiagnosis", "")
        row["Reasoning_output_PrimaryDiagnosis"] = all_data["Reasoning_output"].get(img_name, {}).get("PrimaryDiagnosis", "")
        row["TreatmentRecommend_output_PrimaryDiagnosis"] = all_data["TreatmentRecommend_output"].get(img_name, {}).get("PrimaryDiagnosis", "")

        # 新增 'label' 列，值为图片名去掉 .jpg 后缀
        row["label"] = os.path.splitext(img_name)[0]

        csv_rows.append(row)

    # 创建 Pandas DataFrame
    df = pd.DataFrame(csv_rows)

    # 确保“图像名”是第一列
    cols = ["图像名"] + [col for col in df.columns if col not in ["图像名", "label"]] + ["label"]
    df = df[cols]

    # 将DataFrame写入CSV文件
    df.to_csv(output_csv_path, index=False, encoding='utf-8')

    print(f"数据已成功提取并写入到 '{output_csv_path}'")

# ==================== 仅替换 HighLight() 函数 ====================
def HighLight(INPUT_CSV, OUTPUT_XLSX, TARGET_COLS):
    FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    # 近义词映射表：出现左侧任意词 → 统一成右侧
    NORM_MAP = {
        "dermatographia": "dermatographia",
        "dermatographic": "dermatographia",
        "dermagraphism": "dermatographia",
        "dermographism": "dermatographia",
        "puppp": "puppp",
        "pep": "puppp",  # 如果想让 PEP 也算 PUPPP，就保留；否则删
    }

    def normalize(word: str) -> str:
        return NORM_MAP.get(word, word)

    def extract_words(text: str) -> set:
        if pd.isna(text):
            return set()
        cleaned = re.sub(r"\([^)]*\)", " ", str(text))
        return {normalize(w) for w in re.findall(r"[A-Za-z]+", cleaned.lower())}

    df = pd.read_csv(INPUT_CSV)
    df.to_excel(OUTPUT_XLSX, index=False)

    wb  = load_workbook(OUTPUT_XLSX)
    ws  = wb.active
    col_name_to_letter = {cell.value: cell.column_letter for cell in ws[1]}

    for row_idx in range(2, ws.max_row + 1):
        label_cell  = ws.cell(row=row_idx, column=df.columns.get_loc("label") + 1)
        label_words = {normalize(w) for w in label_cell.value.lower().split("-")}

        for col_name in TARGET_COLS:
            col_letter = col_name_to_letter[col_name]
            cell       = ws[f"{col_letter}{row_idx}"]
            content_words = extract_words(cell.value)

            # 归一化后完全相等即命中
            if content_words & label_words:
                cell.fill = FILL_GREEN

    wb.save(OUTPUT_XLSX)
    print(f"✅ 近义词归一化高亮完成：{OUTPUT_XLSX}")
# ==================== 替换结束 ====================
if __name__=="__main__":
    # ClassificationResultSaveToCSV('./SkinGPT-X-EvaluationResults/Dermnet/test/RAG_output.json',
    #                               './SkinGPT-X-EvaluationResults/Dermnet/test/SkinGPT_output.json',
    #                               './SkinGPT-X-EvaluationResults/Dermnet/test/WebSearch_output.json',
    #                               './SkinGPT-X-EvaluationResults/Dermnet/test/CaseReview_output.json',
    #                               './SkinGPT-X-EvaluationResults/Dermnet/test/Reasoning_output.json',
    #                               './SkinGPT-X-EvaluationResults/Dermnet/test/TreatmentRecommend_output.json',
    #                               './SkinGPT-X-EvaluationResults/Dermnet/test/ClassificationResults.csv')
    HighLight(
        INPUT_CSV="./SkinGPT-X-EvaluationResults/Dermnet/test/ClassificationResults.csv",
        OUTPUT_XLSX="./SkinGPT-X-EvaluationResults/Dermnet/test/ClassificationResults_highlighted.xlsx",
        TARGET_COLS=[
            "RAG_output_诊断评估值",
            "SkinGPT_output_诊断评估值",
            "WebSearch_output_诊断评估值",
            "CaseReview_output_PrimaryDiagnosis",
            "Reasoning_output_PrimaryDiagnosis",
            "TreatmentRecommend_output_PrimaryDiagnosis",
        ])
